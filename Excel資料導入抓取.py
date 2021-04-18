import openpyxl
f = 'sales.xlsx'
wb = openpyxl.load_workbook(f,data_only=True)
#data_only=True此公式可顯示結果跳過值
ws = wb['2020Q1']

print(ws['A1'].value)
print(ws['A2'].value)
print(ws['B2'].value)
print(ws['B3'].value)
print(ws['C5'].value)
print(ws['E3'].value)
print()
print(ws['E3'].column)#欄
print(ws['E3'].row)#列
print(ws['E3'].coordinate)#位置