import openpyxl
# wb = openpyxl.Workbook()#新建
wb = openpyxl.load_workbook('demo3.xlsx')#讀取
# wb.create_sheet('學生表',1)
# wb.create_sheet('出席紀錄',2)

ws = wb.active
# ws.title = 'ChiChi'#下方活頁檔名稱


title = ['Name','Dep']
ws.append(title)
while True:
    stu = list()
    name = input('St Name :')
    if name =='':
        break
    dep = input('Dep :')
    stu.append(name)
    stu.append(dep)
    ws.append(stu)
wb.save('demo3.xlsx')