import openpyxl
wb = openpyxl.Workbook()
wb.create_sheet('學生表',1)
wb.create_sheet('出席紀錄',0)
ws = wb['學生表']
title = ['Name','Dep']
ws.append(title)
while True:
    stu = list()
    name = input('St Name :')
    if name =='':
        break
    dep = input('Dep :')
    stu.append(name)
    stu.append(dep)
    ws.append(stu)
wb.save('demo2.xlsx')